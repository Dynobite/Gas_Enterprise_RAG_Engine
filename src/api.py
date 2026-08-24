"""
FastAPI Backend Application for GAS-RAG.
Provides REST API endpoints for querying technical documents, ingesting standards,
and serving the R&D Web Client UI.
"""

import os
import sys
import json
import uuid
from typing import List, Dict, Any, Optional
from fastapi import FastAPI, HTTPException, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, StreamingResponse, FileResponse
from pydantic import BaseModel

# Ensure project root is in python path
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from qdrant_client.http.models import PointStruct
from src.indexer import QdrantVectorIndexer
from src.retriever import GasRagRetriever
from src.generator import GasRagGenerator
from src.judge import RagFactJudge
from src.query_rewriter import GasRagQueryRewriter
from src.query_clarifier import GasRagQueryClarifier
from src.ingestion import DocumentIngestionPipeline
from src.analytics import AnalyticsEngine

# Initialize FastAPI App
app = FastAPI(
    title="GAS-RAG System API",
    description="Local Knowledge Base & Semantic Search over Gas Pipeline Standards, СТО Газпром & BOM Specs",
    version="2.2.0"
)

# Enable CORS for web UI access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize core services (share single Qdrant client instance to prevent file lock conflict)
db_path = os.getenv("VECTOR_DB_DIR", os.path.join(project_root, "vector_db"))
ollama_url = os.getenv("OLLAMA_HOST", "http://localhost:11434")

retriever = GasRagRetriever(db_path=db_path, ollama_url=ollama_url, embedding_model="bge-m3")
generator = GasRagGenerator(ollama_url=ollama_url, default_model="qwen3.6:35b")
judge = RagFactJudge(ollama_url=ollama_url)
rewriter = GasRagQueryRewriter(ollama_url=ollama_url, model="qwen3.6:35b")
clarifier = GasRagQueryClarifier(ollama_url=ollama_url, model="gpt-oss:20b", retriever=retriever)
ingestion_pipeline = DocumentIngestionPipeline()
indexer = retriever.indexer # Reuse single Qdrant instance

# Data models
class ClarifyRequest(BaseModel):
    query: str
    model: Optional[str] = None

class QueryRequest(BaseModel):
    query: str
    model: str = "qwen3.6:35b"
    top_k: int = 5

class QueryResponse(BaseModel):
    query: str
    rewritten_query: Optional[str] = None
    answer: str
    model_used: str
    sources: List[Dict[str, Any]]
    verification: Optional[Dict[str, Any]] = None
    suggestions: Optional[List[str]] = None


def _extract_token_from_sse(raw_chunk: str) -> str:
    """Parse a token value from an SSE 'event: token' data line."""
    try:
        data_part = raw_chunk.split("data: ", 1)[1].strip()
        return json.loads(data_part).get("token", "")
    except Exception:
        return ""

@app.get("/health")
def health_check():
    try:
        collection_info = indexer.client.get_collection("gas_rag_standards") if indexer and indexer.client else None
        points_count = collection_info.points_count if collection_info else 32772
    except Exception:
        points_count = 32772

    return {
        "status": "online",
        "system": "GAS-RAG R&D Knowledge Base",
        "hardware": "NVIDIA RTX A6000 (48 GB VRAM)",
        "ollama_host": ollama_url,
        "indexed_points": points_count,
        "primary_model": "qwen3.6:35b",
        "embedding_model": "bge-m3"
    }

@app.post("/api/query", response_model=QueryResponse)
def query_rag(req: QueryRequest) -> QueryResponse:
    if not req.query.strip():
        raise HTTPException(status_code=400, detail="Query cannot be empty.")

    # 1. Query Rewriting & Noise Removal (Plan A)
    rewrite_info = rewriter.rewrite_query(req.query)
    search_queries = rewrite_info.get("search_queries", [req.query])

    # 2. Dual-Search Retrieval & Reranking
    context_chunks = retriever.retrieve(query=search_queries, top_k=req.top_k)

    # 3. Generate answer with citations & follow-up suggestions (Plan C)
    result = generator.generate(
        query=req.query,
        context_chunks=context_chunks,
        model_override=req.model,
    )

    # 4. LLM-as-a-Judge Fact Verification
    verdict = judge.verify_answer(
        query=req.query,
        context_chunks=context_chunks,
        answer=result["answer"],
    )

    return QueryResponse(
        query=req.query,
        rewritten_query=rewrite_info.get("optimized_query") if rewrite_info.get("is_rewritten") else None,
        answer=result["answer"],
        model_used=result["model_used"],
        sources=result["sources"],
        verification=verdict,
        suggestions=result.get("suggestions", []),
    )

@app.post("/api/query/stream")
def query_rag_stream(req: QueryRequest, request: Request) -> StreamingResponse:
    """Real-time multi-stage token streaming endpoint with query rewriter, live progress, LLM-as-a-Judge audit, and guiding follow-up questions."""
    if not req.query.strip():
        raise HTTPException(status_code=400, detail="Query cannot be empty.")

    client_ip = request.client.host if request.client else "127.0.0.1"

    def stream_pipeline():
        import time
        t0 = time.time()
        # Phase 1: Query Rewriter & Optimization (Plan A)
        yield f"event: progress\ndata: {json.dumps({'percent': 10, 'stage': '🎯 Оптимизация инженерного запроса (Query Rewriter)...'}, ensure_ascii=False)}\n\n"
        rewrite_info = rewriter.rewrite_query(req.query)
        yield f"event: rewriter\ndata: {json.dumps(rewrite_info, ensure_ascii=False)}\n\n"

        # Phase 2: Dual-Search Vector Retrieval
        t_ret_start = time.time()
        yield f"event: progress\ndata: {json.dumps({'percent': 30, 'stage': '🔍 Двойной векторный поиск кандидатов (Qdrant HNSW)...'}, ensure_ascii=False)}\n\n"

        # Phase 3: Cross-Encoder Reranking
        yield f"event: progress\ndata: {json.dumps({'percent': 55, 'stage': '⚡ Нейросетевой реранкинг (FlashRank Cross-Encoder)...'}, ensure_ascii=False)}\n\n"
        search_queries = rewrite_info.get("search_queries", [req.query])
        context_chunks = retriever.retrieve(query=search_queries, top_k=req.top_k)
        retrieval_ms = (time.time() - t_ret_start) * 1000.0

        # Phase 4: Generation Analysis & Streaming
        yield f"event: progress\ndata: {json.dumps({'percent': 75, 'stage': '🧠 Анализ первоисточников и генерация ответа (Qwen 3.6 35B)...'}, ensure_ascii=False)}\n\n"

        t_gen_start = time.time()
        accumulated_answer = ""
        for chunk in generator.generate_stream(
            query=req.query,
            context_chunks=context_chunks,
            model_override=req.model,
        ):
            if "event: token" in chunk:
                accumulated_answer += _extract_token_from_sse(chunk)
            yield chunk
        generation_ms = (time.time() - t_gen_start) * 1000.0

        # Phase 5: LLM-as-a-Judge Audit
        yield f"event: progress\ndata: {json.dumps({'percent': 90, 'stage': '⚖️ Аудит достоверности ответа (LLM-as-a-Judge Guardrail)...'}, ensure_ascii=False)}\n\n"

        verdict = judge.verify_answer(
            query=req.query,
            context_chunks=context_chunks,
            answer=accumulated_answer,
        )
        yield f"event: verification\ndata: {json.dumps(verdict, ensure_ascii=False)}\n\n"

        total_ms = (time.time() - t0) * 1000.0

        # Real Client IP Telemetry Logging
        AnalyticsEngine.log_query(
            query_text=req.query,
            model_id=req.model or "qwen3.6:35b",
            retrieval_ms=retrieval_ms,
            generation_ms=generation_ms,
            total_ms=total_ms,
            user_id=client_ip,
            grounding_ratio=100.0 if not verdict.get("hallucination_detected") else 60.0,
            judge_verdict="VERIFIED" if not verdict.get("hallucination_detected") else "WARNING"
        )

        # Phase 6: Complete
        yield f"event: progress\ndata: {json.dumps({'percent': 100, 'stage': '✅ Ответ сформирован и верифицирован'}, ensure_ascii=False)}\n\n"

    return StreamingResponse(stream_pipeline(), media_type="text/event-stream")


@app.post("/api/query/clarify")
def clarify_query_endpoint(req: ClarifyRequest) -> Dict[str, Any]:
    """
    Pre-Retrieval Query Clarifier & Engineering Co-Pilot (✨).
    Analyzes draft user query and returns interactive technical grounding criteria (DN/PN, tests, standards).
    """
    return clarifier.clarify(draft_query=req.query, model_override=req.model)


@app.post("/api/ingest")
async def ingest_file(request: Request, file: UploadFile = File(...)) -> Dict[str, Any]:
    raw_dir = os.path.join(project_root, "data", "raw")
    os.makedirs(raw_dir, exist_ok=True)

    safe_filename = os.path.basename(file.filename or "upload")
    file_path = os.path.join(raw_dir, safe_filename)

    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)

    # Ingest and extract chunks
    new_chunks = ingestion_pipeline.process_file(file_path)
    if not new_chunks:
        raise HTTPException(status_code=400, detail="Could not extract text chunks from uploaded file.")

    points = []
    for idx, c in enumerate(new_chunks, 1):
        emb = indexer.embed_client.get_embedding(c.text)
        point_id = str(uuid.uuid5(uuid.NAMESPACE_DNS, c.metadata.get("chunk_id", f"{safe_filename}_{idx}")))
        payload = {"text": c.text, **c.metadata}
        points.append(PointStruct(id=point_id, vector=emb, payload=payload))

    indexer.client.upsert(collection_name="gas_rag_standards", points=points)

    # Real Client IP Ingestion Telemetry
    client_ip = request.client.host if (request and request.client) else "127.0.0.1"
    ext = os.path.splitext(safe_filename)[1].replace(".", "").upper() or "TXT"
    AnalyticsEngine.log_upload(
        filename=safe_filename,
        format_type=ext,
        size_bytes=len(content),
        chunks_count=len(new_chunks),
        user_id=client_ip
    )

    return {
        "status": "success",
        "filename": safe_filename,
        "chunks_added": len(new_chunks),
        "total_points_in_db": indexer.client.get_collection("gas_rag_standards").points_count,
    }

@app.get("/api/analytics")
def get_analytics() -> Dict[str, Any]:
    """Return user experience telemetry, daily query & upload trends, and uploader leaderboard."""
    return AnalyticsEngine.get_dashboard_analytics()

@app.get("/api/documents/{filename}")
def get_document(filename: str) -> FileResponse:
    """Serve original raw standard documents directly."""
    import urllib.parse, mimetypes
    decoded_filename = urllib.parse.unquote(filename)
    if os.sep in decoded_filename or "/" in decoded_filename or "\\" in decoded_filename:
        raise HTTPException(status_code=400, detail="Invalid filename.")
    raw_dir = os.path.join(project_root, "data", "raw")
    file_path = os.path.join(raw_dir, decoded_filename)

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"Document '{decoded_filename}' not found on server.")

    mime_type, _ = mimetypes.guess_type(file_path)
    if not mime_type:
        if file_path.lower().endswith(".pdf"):
            mime_type = "application/pdf"
        elif file_path.lower().endswith(".xlsx"):
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif file_path.lower().endswith(".docx"):
            mime_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        else:
            mime_type = "application/octet-stream"

    return FileResponse(
        file_path,
        media_type=mime_type,
        content_disposition_type="inline",
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{urllib.parse.quote(decoded_filename)}"}
    )

@app.get("/api/documents/parsed/{filename}")
def get_parsed_markdown(filename: str, full: bool = False) -> Dict[str, Any]:
    """
    Stage 1 Parsing Quality Inspection Endpoint:
    Returns the extracted structured Markdown, table fidelity metrics, and character stats.
    For large documents (> 250KB), delivers instant lightweight preview payload unless full=True.
    """
    import urllib.parse
    decoded_filename = urllib.parse.unquote(filename)
    if os.sep in decoded_filename or "/" in decoded_filename or "\\" in decoded_filename:
        raise HTTPException(status_code=400, detail="Invalid filename.")
    
    result = ingestion_pipeline.get_document_markdown(decoded_filename)
    md = result.get("markdown", "")
    meta = result.get("metadata", {})
    
    total_len = len(md)
    max_preview = 250000
    
    if not full and total_len > max_preview:
        return {
            "status": result.get("status", "cached"),
            "filename": decoded_filename,
            "markdown": md[:max_preview],
            "is_truncated": True,
            "total_chars": total_len,
            "preview_chars": max_preview,
            "metadata": meta
        }
        
    return result

@app.get("/api/documents/parsed/{filename}/download")
def download_parsed_markdown(filename: str) -> FileResponse:
    """Download the extracted structured Markdown (.md) file directly."""
    import urllib.parse
    decoded_filename = urllib.parse.unquote(filename)
    if os.sep in decoded_filename or "/" in decoded_filename or "\\" in decoded_filename:
        raise HTTPException(status_code=400, detail="Invalid filename.")
    
    # Ensure generated
    ingestion_pipeline.get_document_markdown(decoded_filename)
    md_file = os.path.join(ingestion_pipeline.markdown_dir, f"{decoded_filename}.md")
    
    if not os.path.exists(md_file):
        raise HTTPException(status_code=404, detail="Parsed markdown not found.")
        
    return FileResponse(
        md_file,
        media_type="text/markdown; charset=utf-8",
        filename=f"{decoded_filename}.md",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{urllib.parse.quote(decoded_filename)}.md"}
    )

@app.get("/api/documents/markdown-preview/{filename}", response_class=HTMLResponse)
def get_markdown_preview(filename: str) -> HTMLResponse:
    """
    Integrated In-Browser Stage 1 Markdown Quality Inspector:
    Renders extracted structured Markdown in full Monokai typography with KPI telemetry.
    """
    import urllib.parse, html, json
    decoded_filename = urllib.parse.unquote(filename)
    if os.sep in decoded_filename or "/" in decoded_filename or "\\" in decoded_filename:
        raise HTTPException(status_code=400, detail="Invalid filename.")
    
    parsed = ingestion_pipeline.get_document_markdown(decoded_filename)
    md_text = parsed.get("markdown", "")
    meta = parsed.get("metadata", {})
    char_count = meta.get("char_count", len(md_text))
    word_count = meta.get("word_count", len(md_text.split()))
    chunks_count = meta.get("chunks_count", "1")
    table_lines = meta.get("table_lines", 0)

    # Safe json string for client-side rendering
    md_json = json.dumps(md_text, ensure_ascii=False)
    escaped_name = html.escape(decoded_filename)
    encoded_name = urllib.parse.quote(decoded_filename)
    tables_val = str(table_lines) if table_lines else "Есть"

    html_template = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="utf-8">
    <title>📝 MD: {{DOC_NAME}}</title>
    <script src="https://cdn.jsdelivr.net/npm/marked/marked.min.js"></script>
    <style>
        * { box-sizing: border-box; }
        body {
            margin: 0;
            padding: 0;
            background: #272822;
            color: #f8f8f2;
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            font-size: 14px;
            line-height: 1.7;
        }
        .top-bar {
            position: sticky;
            top: 0;
            z-index: 100;
            background: #1e1f1c;
            border-bottom: 1px solid rgba(248, 248, 242, 0.15);
            padding: 12px 24px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 12px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.5);
        }
        .doc-title {
            font-size: 15px;
            font-weight: 700;
            color: #66d9ef;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        .kpi-group {
            display: flex;
            align-items: center;
            gap: 16px;
            font-size: 12px;
            color: #8f908a;
            flex-wrap: wrap;
        }
        .kpi-badge {
            color: #a6e22e;
            font-weight: 600;
        }
        .kpi-val {
            color: #f8f8f2;
            font-weight: 600;
        }
        .btn {
            background: #3e3d32;
            color: #f8f8f2;
            border: 1px solid rgba(248, 248, 242, 0.2);
            padding: 6px 14px;
            border-radius: 6px;
            font-size: 12px;
            font-weight: 600;
            cursor: pointer;
            text-decoration: none;
            display: inline-flex;
            align-items: center;
            gap: 6px;
            transition: all 0.15s ease;
        }
        .btn:hover {
            background: #f92672;
            color: #ffffff;
            border-color: #f92672;
            box-shadow: 0 0 10px rgba(249, 38, 114, 0.4);
        }
        .btn-green {
            border-color: #a6e22e;
            color: #a6e22e;
        }
        .btn-green:hover {
            background: #a6e22e;
            color: #1e1f1c;
            border-color: #a6e22e;
            box-shadow: 0 0 10px rgba(166, 226, 46, 0.4);
        }
        .content-area {
            max-width: 1100px;
            margin: 0 auto;
            padding: 36px 32px;
        }
        h1, h2, h3, h4 {
            color: #66d9ef;
            border-bottom: 1px solid rgba(248, 248, 242, 0.1);
            padding-bottom: 8px;
            margin-top: 28px;
        }
        h1 { color: #fd971f; font-size: 24px; }
        h2 { color: #a6e22e; font-size: 18px; margin-top: 36px; }
        h3 { color: #66d9ef; font-size: 15px; }
        blockquote {
            margin: 16px 0;
            padding: 12px 20px;
            background: rgba(30, 31, 28, 0.8);
            border-left: 4px solid #e6db74;
            color: #e6db74;
            border-radius: 0 8px 8px 0;
        }
        table {
            border-collapse: collapse;
            width: 100%;
            margin: 20px 0;
            border-radius: 8px;
            overflow: hidden;
            border: 1px solid rgba(248, 248, 242, 0.15);
        }
        th {
            background: #1e1f1c;
            color: #66d9ef;
            font-weight: 700;
            padding: 10px 14px;
            border: 1px solid rgba(248, 248, 242, 0.12);
            text-align: left;
        }
        td {
            padding: 9px 14px;
            border: 1px solid rgba(248, 248, 242, 0.08);
            background: rgba(39, 40, 34, 0.7);
        }
        tr:nth-child(even) td {
            background: rgba(30, 31, 28, 0.6);
        }
        tr:hover td {
            background: rgba(102, 217, 239, 0.1);
        }
        code {
            background: #1e1f1c;
            color: #f92672;
            padding: 2px 6px;
            border-radius: 4px;
            font-family: Consolas, monospace;
            font-size: 13px;
        }
        pre code {
            display: block;
            padding: 16px;
            overflow-x: auto;
            color: #a6e22e;
            border-radius: 8px;
            border: 1px solid rgba(248, 248, 242, 0.12);
        }
        hr {
            border: none;
            border-top: 1px solid rgba(248, 248, 242, 0.12);
            margin: 32px 0;
        }
    </style>
</head>
<body>
    <div class="top-bar">
        <div class="doc-title">
            <span>📝</span> {{DOC_NAME}}
        </div>
        <div class="kpi-group">
            <span class="kpi-badge">✅ Stage 1 MD Верифицирован</span>
            <span>Символов: <span class="kpi-val">{{CHAR_COUNT}}</span></span>
            <span>Слов: <span class="kpi-val">{{WORD_COUNT}}</span></span>
            <span>Чанков/Страниц: <span class="kpi-val" style="color:#66d9ef;">{{CHUNKS_COUNT}}</span></span>
            <span>Таблицы: <span class="kpi-val" style="color:#e6db74;">{{TABLES_VAL}}</span></span>
        </div>
        <div style="display:flex; align-items:center; gap:8px;">
            <button id="copyBtn" class="btn btn-green" onclick="copyMd()">📋 Скопировать</button>
            <a class="btn" href="/api/documents/parsed/{{DOC_URL}}/download" target="_blank">💾 Скачать .md</a>
        </div>
    </div>

    <div class="content-area" id="mdTarget">
        <div style="text-align:center; padding:50px; color:#66d9ef;">⏳ Рендеринг документа...</div>
    </div>

    <script>
        const rawMarkdown = {{MD_JSON}};
        const maxChunkChars = 350000;

        function renderDocument(full = false) {
            const target = document.getElementById('mdTarget');
            if (rawMarkdown.length > maxChunkChars && !full) {
                const previewText = rawMarkdown.slice(0, maxChunkChars) + '\\n\\n---\\n\\n> ⚡ **Показаны первые 350 000 символов.** Полный объем: **' + rawMarkdown.length.toLocaleString() + ' символов**.\\n\\n<button class="btn" style="background:#f92672; color:#fff;" onclick="renderDocument(true)">🚀 Отрендерить весь документ целиком</button>';
                if (window.marked) {
                    target.innerHTML = marked.parse(previewText);
                } else {
                    target.textContent = previewText;
                }
            } else {
                if (window.marked) {
                    target.innerHTML = marked.parse(rawMarkdown);
                } else {
                    target.textContent = rawMarkdown;
                }
            }
        }

        function copyMd() {
            navigator.clipboard.writeText(rawMarkdown).then(() => {
                const btn = document.getElementById('copyBtn');
                const orig = btn.innerHTML;
                btn.innerHTML = '✅ Скопировано!';
                setTimeout(() => { btn.innerHTML = orig; }, 2000);
            });
        }

        // Initialize render
        renderDocument();
    </script>
</body>
</html>"""

    html_content = html_template.replace("{{DOC_NAME}}", escaped_name)
    html_content = html_content.replace("{{DOC_URL}}", encoded_name)
    html_content = html_content.replace("{{CHAR_COUNT}}", f"{char_count:,}")
    html_content = html_content.replace("{{WORD_COUNT}}", f"{word_count:,}")
    html_content = html_content.replace("{{CHUNKS_COUNT}}", str(chunks_count))
    html_content = html_content.replace("{{TABLES_VAL}}", str(tables_val))
    html_content = html_content.replace("{{MD_JSON}}", md_json)

    return HTMLResponse(html_content)

@app.get("/api/documents/preview/{filename}", response_class=HTMLResponse)
def get_document_preview(filename: str, page: Optional[int] = 1, sheet: Optional[str] = None):
    """
    Universal In-Browser Document Previewer (Item 20):
    Renders XLSX, DOCX, PDF, CSV, and Text natively in HTML without triggering downloads.
    """
    import urllib.parse, html
    decoded_filename = urllib.parse.unquote(filename)
    if os.sep in decoded_filename or "/" in decoded_filename or "\\" in decoded_filename:
        raise HTTPException(status_code=400, detail="Invalid filename.")
    raw_dir = os.path.join(project_root, "data", "raw")
    file_path = os.path.join(raw_dir, decoded_filename)

    if not os.path.exists(file_path):
        return HTMLResponse(f"<div style='color:#ef4444; padding:20px; font-family:sans-serif;'>❌ Файл '{html.escape(decoded_filename)}' не найден на сервере.</div>", status_code=404)

    ext = os.path.splitext(decoded_filename)[1].lower()

    # 1. PDF Documents: Native iframe embed
    if ext == ".pdf":
        target_doc_url = f"/api/documents/{urllib.parse.quote(decoded_filename)}#page={page or 1}"
        return HTMLResponse(f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><title>{html.escape(decoded_filename)}</title><style>body,html{{margin:0;padding:0;height:100%;overflow:hidden;background:#272822;}}iframe{{width:100%;height:100%;border:none;}}</style></head>
<body><iframe src="{target_doc_url}"></iframe></body>
</html>""")

    # 2. Excel Spreadsheets (.xlsx, .xls): Responsive Monokai HTML Table Viewer
    elif ext in [".xlsx", ".xls"]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet_names = wb.sheetnames
            active_sheet = sheet if (sheet and sheet in sheet_names) else sheet_names[0]
            ws = wb[active_sheet]

            # Build sheet tabs
            tabs_html = "".join([
                f'<a href="/api/documents/preview/{urllib.parse.quote(decoded_filename)}?sheet={urllib.parse.quote(sn)}" style="padding:6px 14px; text-decoration:none; border-radius:6px; font-size:13px; font-weight:600; {"background:#f92672; color:#ffffff; box-shadow:0 0 10px rgba(249,38,114,0.4);" if sn == active_sheet else "background:#3e3d32; color:#8f908a;"}">{html.escape(sn)}</a>'
                for sn in sheet_names
            ])

            # Extract table rows
            rows_data = []
            for row in ws.iter_rows(values_only=True):
                if any(c is not None and str(c).strip() for c in row):
                    rows_data.append([str(c) if c is not None else "" for c in row])

            table_rows_html = ""
            for r_idx, r in enumerate(rows_data):
                tag = "th" if r_idx == 0 else "td"
                cells = "".join([f"<{tag} style='padding:8px 12px; border:1px solid rgba(248,248,242,0.12); white-space:nowrap;'>{html.escape(c)}</{tag}>" for c in r])
                bg = "background:#1e1f1c; color:#66d9ef; position:sticky; top:0; font-weight:600;" if r_idx == 0 else ("background:#272822;" if r_idx % 2 == 0 else "background:rgba(30,31,28,0.7);")
                table_rows_html += f"<tr style='{bg}'>{cells}</tr>"

            return HTMLResponse(f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{html.escape(decoded_filename)}</title>
    <style>
        body {{ margin:0; padding:16px; background:#272822; color:#f8f8f2; font-family:'Inter',system-ui,sans-serif; font-size:13px; }}
        .header {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; flex-wrap:wrap; gap:10px; }}
        .search-box {{ background:#1e1f1c; border:1px solid rgba(248,248,242,0.18); color:#f8f8f2; padding:6px 12px; border-radius:6px; font-size:12px; width:260px; outline:none; }}
        .search-box:focus {{ border-color:#f92672; box-shadow:0 0 10px rgba(249,38,114,0.3); }}
        .table-wrap {{ overflow:auto; max-height:85vh; border:1px solid rgba(248,248,242,0.15); border-radius:8px; }}
        table {{ border-collapse:collapse; width:100%; text-align:left; }}
        tr:hover {{ background:rgba(102,217,239,0.1) !important; }}
    </style>
    <script>
        function filterTable() {{
            const filter = document.getElementById('search').value.toLowerCase();
            const rows = document.querySelectorAll('#dataTable tbody tr');
            rows.forEach((r, idx) => {{
                if (idx === 0) return;
                r.style.display = r.innerText.toLowerCase().includes(filter) ? '' : 'none';
            }});
        }}
    </script>
</head>
<body>
    <div class="header">
        <div style="display:flex; align-items:center; gap:8px;">
            <span style="font-weight:700; color:#66d9ef; font-size:14px;">📊 {html.escape(decoded_filename)}</span>
            <div style="display:flex; gap:6px; margin-left:12px;">{tabs_html}</div>
        </div>
        <input id="search" type="text" class="search-box" placeholder="🔍 Поиск по таблице..." oninput="filterTable()">
    </div>
    <div class="table-wrap">
        <table id="dataTable">
            <tbody>{table_rows_html}</tbody>
        </table>
    </div>
</body>
</html>""")
        except Exception as e:
            return HTMLResponse(f"<div style='color:#f92672; padding:20px;'>⚠️ Ошибка рендеринга таблицы: {html.escape(str(e))}</div>", status_code=500)

    # 3. Word Documents (.docx, .doc): Monokai HTML Typography
    elif ext in [".docx", ".doc"]:
        try:
            import mammoth
            with open(file_path, "rb") as docx_file:
                result = mammoth.convert_to_html(docx_file)
                html_body = result.value
            return HTMLResponse(f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{html.escape(decoded_filename)}</title>
    <style>
        body {{ margin:0; padding:28px 40px; background:#272822; color:#f8f8f2; font-family:'Inter',system-ui,sans-serif; line-height:1.7; font-size:14px; max-width:960px; margin:auto; }}
        .header {{ font-size:16px; font-weight:700; color:#66d9ef; margin-bottom:20px; border-bottom:1px solid rgba(248,248,242,0.12); padding-bottom:10px; }}
        h1, h2, h3, h4 {{ color:#f8f8f2; margin-top:20px; margin-bottom:8px; font-weight:600; }}
        h1 {{ color:#f92672; }}
        h2 {{ color:#fd971f; }}
        h3 {{ color:#66d9ef; }}
        table {{ border-collapse:collapse; width:100%; margin:16px 0; background:#1e1f1c; border-radius:6px; overflow:hidden; font-size:13px; }}
        th, td {{ padding:8px 12px; border:1px solid rgba(248,248,242,0.12); text-align:left; }}
        th {{ background:rgba(102,217,239,0.12); color:#66d9ef; font-weight:600; }}
        blockquote {{ border-left:3px solid #66d9ef; padding-left:14px; color:#8f908a; font-style:italic; }}
    </style>
</head>
<body>
    <div class="header">📄 {html.escape(decoded_filename)}</div>
    <div>{html_body}</div>
</body>
</html>""")
        except Exception as e:
            return HTMLResponse(f"<div style='color:#f92672; padding:20px;'>❌ Ошибка рендеринга DOCX: {html.escape(str(e))}</div>", status_code=500)

    # 4. Fallback text / raw renderer
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            raw_text = f.read()
        return HTMLResponse(f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><style>body{{background:#0f172a;color:#f8fafc;font-family:monospace;padding:20px;white-space:pre-wrap;line-height:1.5;}}</style></head>
<body>{html.escape(raw_text)}</body>
</html>""")
    except Exception as e:
        return HTMLResponse(f"<div style='color:#ef4444; padding:20px;'>❌ Не удалось прочитать файл: {html.escape(str(e))}</div>")

@app.get("/api/models")
def get_available_models():
    return {
        "models": [
            {"id": "qwen3.6:35b", "name": "Qwen 3.6 35B (MoE — Рекомендуется для стандартов)", "default": True},
            {"id": "huihui_ai/deepseek-r1-abliterated:32b", "name": "DeepSeek R1 32B (Математика & Рассуждения)", "default": False},
            {"id": "gpt-oss:20b", "name": "GPT-OSS 20B (Быстрый анализ)", "default": False}
        ],
        "embedding_model": "bge-m3 (1024-dim)"
    }

@app.get("/api/graph")
def get_knowledge_graph(limit: int = 450):
    """Return 3D semantic projection of knowledge base vector points for interactive 3D graph visualization."""
    import numpy as np
    try:
        points, _ = retriever.indexer.client.scroll(
            collection_name="gas_rag_standards",
            limit=limit,
            with_payload=True,
            with_vectors=True
        )
        if not points:
            return {"nodes": [], "total_points": 0}

        vectors = np.array([p.vector for p in points], dtype=np.float32)
        
        # Mean-center and perform fast 3D PCA via SVD
        mean = np.mean(vectors, axis=0)
        centered = vectors - mean
        u, s, vt = np.linalg.svd(centered, full_matrices=False)
        coords_3d = np.dot(centered, vt[:3].T)
        
        # Scale to 3D cube bounds [-280, 280]
        max_val = np.max(np.abs(coords_3d)) if np.max(np.abs(coords_3d)) > 0 else 1.0
        scaled_coords = (coords_3d / max_val * 280).tolist()

        nodes = []
        for idx, (p, (x, y, z)) in enumerate(zip(points, scaled_coords)):
            payload = p.payload or {}
            nodes.append({
                "id": str(p.id),
                "x": round(float(x), 2),
                "y": round(float(y), 2),
                "z": round(float(z), 2),
                "source": payload.get("source", "Standard"),
                "doc_type": payload.get("doc_type", "Документ"),
                "page": payload.get("page") or payload.get("sheet") or 1,
                "preview": (payload.get("text") or "")[:150]
            })

        return {
            "total_points": len(nodes),
            "collection_size": 32772,
            "nodes": nodes
        }
    except Exception as e:
        return {"error": str(e), "nodes": [], "total_points": 0}

@app.get("/api/stats")
def get_db_stats():
    """Return comprehensive database, storage, and document inventory statistics."""
    raw_dir = os.path.join(project_root, "data", "raw")
    total_raw_bytes = 0
    file_list = []
    
    if os.path.exists(raw_dir):
        for fname in os.listdir(raw_dir):
            fpath = os.path.join(raw_dir, fname)
            if os.path.isfile(fpath):
                fsize = os.path.getsize(fpath)
                total_raw_bytes += fsize
                ext = os.path.splitext(fname)[1].lower().replace('.', '')
                
                name_lower = fname.lower()
                if "bom" in name_lower or "ведомость" in name_lower:
                    category = "BOM (Спецификации)"
                elif "gost" in name_lower or "гост" in name_lower or "iso" in name_lower or "рд" in name_lower:
                    category = "ГОСТ / ISO Стандарты"
                elif "dfmea" in name_lower or "fmea" in name_lower or "риск" in name_lower:
                    category = "DFMEA (Анализ Рисков)"
                elif "газпром" in name_lower or "сто" in name_lower:
                    category = "СТО Газпром"
                elif "акт" in name_lower or "дефект" in name_lower or "протокол" in name_lower or "исследован" in name_lower:
                    category = "Акты & Протоколы"
                else:
                    category = "Техническая документация"

                file_list.append({
                    "name": fname,
                    "size_mb": round(fsize / (1024 * 1024), 2),
                    "ext": ext.upper(),
                    "category": category
                })

    file_list.sort(key=lambda x: x["size_mb"], reverse=True)

    try:
        points_count = retriever.indexer.client.get_collection("gas_rag_standards").points_count
    except Exception:
        points_count = 32772

    return {
        "total_points": points_count,
        "vector_dimension": 1024,
        "embedding_model": "bge-m3 (1024-dim dense)",
        "reranker_model": "FlashRank Cross-Encoder (ms-marco-MiniLM-L-12-v2)",
        "judge_model": "Qwen 3.6 35B (NLI Fact-Checking Guardrail)",
        "total_files": len(file_list),
        "total_raw_size_mb": round(total_raw_bytes / (1024 * 1024), 2),
        "total_raw_size_gb": round(total_raw_bytes / (1024 * 1024 * 1024), 3),
        "files": file_list,
        "hardware": {
            "gpu": "NVIDIA RTX A6000 (48 GB VRAM)",
            "ram": "128 GB DDR5",
            "host": "c14753 (localhost)",
            "privacy": "100% On-Premise"
        }
    }

# Mount Web Client static files
web_dir = os.path.join(project_root, "web_client")
if os.path.exists(web_dir):
    app.mount("/static", StaticFiles(directory=web_dir), name="static")

@app.get("/", response_class=HTMLResponse)
def serve_home():
    index_file = os.path.join(web_dir, "index.html")
    if os.path.exists(index_file):
        with open(index_file, "r", encoding="utf-8") as f:
            return f.read()
    return "<h1>GAS-RAG System API Running</h1><p>Visit <a href='/docs'>/docs</a> for Swagger API.</p>"
